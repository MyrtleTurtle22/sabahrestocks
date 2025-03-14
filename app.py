from flask import Flask, request, send_file
from flask_cors import CORS
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

@app.route('/upload', methods=['POST'])
def upload_files():
    whinventory = request.files['file1']
    storeinventory = request.files['file2']

    dfwh = pd.read_csv(whinventory)
    dfstore = pd.read_csv(storeinventory)

    if "Available" in dfwh.columns:
        dfwh = dfwh.rename(columns={"Available": "WH"})
    if "Available" in dfstore.columns:
        dfstore = dfstore.rename(columns={"Available": "Store"})

    common_cols = set(dfwh.columns) & set(dfstore.columns) - {"Shopify ID", "WH", "Store"}

    dfstore = dfstore.drop(columns=common_cols)

    merged_df = dfwh.merge(dfstore, on="Shopify ID", how="outer")

    merged_df.drop(columns=["SKU", "Shopify ID", "Vendor"], errors='ignore', inplace=True)

    sort_columns = ["Product Type", "Product", "Variant"]
    merged_df.sort_values(by=[col for col in sort_columns if col in merged_df.columns], inplace=True)

    merged_df["SEND"] = ""

    # Reorder columns
    column_order = ["Product Type", "Product", "Variant", "WH", "Store", "Sales", "SEND"]
    merged_df = merged_df[column_order]

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active

    headers = list(merged_df.columns)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, size=14)

        if header == "SEND":
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.font = Font(bold=True, size=16)
        else:
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

    # Add data rows
    for row in merged_df.itertuples(index=False, name=None):
        ws.append(row)

    # Apply thin black borders to all cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    ws.freeze_panes = "A2"

    # Enable filters
    ws.auto_filter.ref = ws.dimensions

    # Adjust column widths
    column_widths = {
        "Product Type": 22,  # Approx. 250 px
        "Product": 40,  # Approx. 450 px
        "Variant": 14,  # Approx. 115 px
        "SEND": 11  # Approx. 120 px
    }
    
    for col_num, header in enumerate(headers, 1):
        if header in column_widths:
            ws.column_dimensions[get_column_letter(col_num)].width = column_widths[header]


    wb.save(output)
    output.seek(0)

    return send_file(output, 
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                     as_attachment=True, 
                     download_name="[Store] Restock [Date].xlsx")

if __name__ == '__main__':
    app.run(debug=True)
